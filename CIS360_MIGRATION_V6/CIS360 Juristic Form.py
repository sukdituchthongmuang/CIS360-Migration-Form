import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side
from sqlalchemy import create_engine
from datetime import datetime

# เชื่อมต่อฐานข้อมูล
DB_USERNAME = "pondttt"
DB_PASSWORD = "2aJ998we~"
DB_NAME = "Customer_UAT"
DB_HOST = "pgsql-sku-az-sea-dev.postgres.database.azure.com"
DB_PORT = "5432"

# กำหนดการจัดวางข้อความตรงกลาง
alignment_center = Alignment(horizontal='center', vertical='center')

# กำหนดกรอบของข้อมูล
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# โหลดไฟล์ Excel ที่มีอยู่
file_path = r'E:/CIS/CIS360 Migration Form V6.xlsx'
wb = load_workbook(file_path)
ws = wb['Juristic Form']

# สร้างการเชื่อมต่อกับฐานข้อมูลด้วยด้วย SQLAlchemy
connection_string = f"postgresql://{DB_USERNAME}:{DB_PASSWORD}@{DB_HOST}:{DB_PORT}/{DB_NAME}"
engine = create_engine(connection_string)

# ดีงค่าจากวิวว์
query = """
SELECT * FROM juristic_form_v6
"""

customer_df = pd.read_sql(query, engine)

# เลือก 10% ของข้อมูล
num_rows = int(len(customer_df) * 0.1)
customer_df = customer_df.head(num_rows).reset_index(drop=True)

# เพิ่มข้อมูลจากฐานข้อมูลลงใน Excel ตั้งแต่แถวที่ 9 เป็นต้นไป
start_row = 9
for idx, row in customer_df.iterrows():
    cus_created_date = row['cus_created_date'].replace(tzinfo=None)
    cus_updated_date = row['cus_updated_date'].replace(tzinfo=None)
    cus_tax_no = str(row['cus_tax_no']) if row['cus_tax_no'] else ''

    ws.cell(row=start_row + idx, column=4, value=str(row['cus_id'])).border = thin_border # reference_id
    ws.cell(row=start_row + idx, column=3, value="ALLKONS_AMRP").border = thin_border # platform_code
    ws.cell(row=start_row + idx, column=7, value="OFFICE").border = thin_border # customer_profile_type
    ws.cell(row=start_row + idx, column=6, value=str(row['cus_no'])).border = thin_border # customer_code
    ws.cell(row=start_row + idx, column=8, value=str(row['cus_company_name'])).border = thin_border # juristic_name
    ws.cell(row=start_row + idx, column=12, value=str(row['cus_tax_no'])).border = thin_border # tax_id
    ws.cell(row=start_row + idx, column=13, value=str(row['cus_branch'])).border = thin_border # branch_number
    ws.cell(row=start_row + idx, column=30, value=str(row['cus_tel'])).border = thin_border # contact_phone_number
    ws.cell(row=start_row + idx, column=32, value=str(row['cus_adress'])).border = thin_border # address_info
    ws.cell(row=start_row + idx, column=21, value=cus_created_date).border = thin_border # create_at
    ws.cell(row=start_row + idx, column=23, value=cus_updated_date).border = thin_border # update_at
    ws.cell(row=start_row + idx, column=17, value="FALSE").border = thin_border # is_dopa
    ws.cell(row=start_row + idx, column=18, value="FALSE").border = thin_border # is_dbd
    ws.cell(row=start_row + idx, column=19, value="APPROVE").border = thin_border # kyc_status
    ws.cell(row=start_row + idx, column=20, value="TRUE").border = thin_border # active_status
    ws.cell(row=start_row + idx, column=9, value=str(row['ct_name'])).border = thin_border # Juristic_type
    ws.cell(row=start_row + idx, column=34, value=str(row['ca_loc_province'])).border = thin_border # province
    ws.cell(row=start_row + idx, column=35, value=str(row['ca_loc_district'])).border = thin_border # district
    ws.cell(row=start_row + idx, column=36, value=str(row['ca_loc_data'])).border = thin_border # sub_district
    ws.cell(row=start_row + idx, column=37, value=str(row['ca_postcode'])).border = thin_border # zipcode
    ws.cell(row=start_row + idx, column=39, value=str(row['ca_loc_province_name'])).border = thin_border # province_name
    ws.cell(row=start_row + idx, column=40, value=str(row['ca_loc_district_name'])).border = thin_border # district_name
    ws.cell(row=start_row + idx, column=41, value=str(row['ca_loc_sub_district_name'])).border = thin_border # sub_district_name
    ws.cell(row=start_row + idx, column=42, value=str(row['ca_postcode'])).border = thin_border # zipcode_name
    ws.cell(row=start_row + idx, column=43, value=str(row['ca_latitude'])).border = thin_border # latitude
    ws.cell(row=start_row + idx, column=44, value=str(row['ca_longitude'])).border = thin_border # longitude

    if ws.cell(row=start_row + idx, column=12).value == 'None': # tax_id
        ws.cell(row=start_row + idx, column=12, value='').border = thin_border
        
    if ws.cell(row=start_row + idx, column=9).value == 'Organization':
        ws.cell(row=start_row + idx, column=9, value="LIMITED_COMPANY").border = thin_border # juristic_type
    elif ws.cell(row=start_row + idx, column=9).value == 'Person':
        ws.cell(row=start_row + idx, column=9, value="Personal").border = thin_border
    else:
        ws.cell(row=start_row + idx, column=9, value="").border = thin_border 
        
    if ws.cell(row=start_row + idx, column=13).value == '00000':
        ws.cell(row=start_row + idx, column=11, value="HEAD_OFFICE").border = thin_border # organize_type
    elif ws.cell(row=start_row + idx, column=13).value == '':
        ws.cell(row=start_row + idx, column=11, value="").border = thin_border
    else:
        ws.cell(row=start_row + idx, column=11, value="BRANCH").border = thin_border
        
    if ws.cell(row=start_row + idx, column=30).value == 'None': # contact_phone_number
        ws.cell(row=start_row + idx, column=30, value='').border = thin_border 

# เพิ่มเส้นกรอบให้กับเซลล์ที่ไม่ได้เติมค่า
for row in range(start_row, start_row + len(customer_df)):
    for col in range(1, 46):
        cell = ws.cell(row=row, column=col)
        if cell.value is None:
            cell.border = thin_border

# บันทึกไฟล์ Excel
wb.save(file_path)
print(f'บันทึกไฟล์ Excel ที่: {file_path} เสร็จสิ้น พร้อมข้อมูลจากฐานข้อมูล')
