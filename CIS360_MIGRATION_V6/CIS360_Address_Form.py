import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side
from sqlalchemy import create_engine
from datetime import datetime

# ข้อมูลการเชื่อมต่อ
DB_USERNAME = "pondttt"
DB_PASSWORD = "2aJ998we~"
DB_NAME = "Customer_UAT"
DB_HOST = "pgsql-sku-az-sea-dev.postgres.database.azure.com"
DB_PORT = "5432"

# กำหนดการจัดวางข้อความตรงกลาง
alignment_center = Alignment(horizontal='center', vertical='center')

# กำหนดกรอบของข้อมูล
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# โหลดไฟล์ Excel ที่มีอยู่
file_path = r'E:/CIS/CIS360 Migration Form V6.xlsx'
wb = load_workbook(file_path)
ws = wb['Address Form']

# สร้างการเชื่อมต่อด้วย SQLAlchemy
connection_string = f"postgresql://{DB_USERNAME}:{DB_PASSWORD}@{DB_HOST}:{DB_PORT}/{DB_NAME}"
engine = create_engine(connection_string)

# สร้าง DataFrame จากฐานข้อมูล
query = """
SELECT * FROM address_form_v6
"""

customer_df = pd.read_sql(query, engine)

# เพิ่มข้อมูลจากฐานข้อมูลลงใน Excel ตั้งแต่แถวที่ 9
start_row = 9
for idx, row in customer_df.iterrows():
    cus_created_date = row['ca_created_date'].replace(tzinfo=None)
    cus_updated_date = row['ca_updated_date'].replace(tzinfo=None)
    cus_tax_no = str(row['cus_tax_no']) if row['cus_tax_no'] else ''
    
    ws.cell(row=start_row + idx, column=3, value="ALLKONS_AMRP").border = thin_border
    ws.cell(row=start_row + idx, column=4, value=str(row['ca_id'])).border = thin_border
    ws.cell(row=start_row + idx, column=5, value="ALLKONS_AMRP").border = thin_border
    ws.cell(row=start_row + idx, column=6, value="OFFICE").border = thin_border
    ws.cell(row=start_row + idx, column=7, value=str(row['ca_cus_id'])).border = thin_border
    ws.cell(row=start_row + idx, column=9, value=str(row['ctl_name'])).border = thin_border
    ws.cell(row=start_row + idx, column=10, value=str(row['ca_name'])).border = thin_border
    ws.cell(row=start_row + idx, column=11, value=str(row['ca_detail_address'])).border = thin_border
    ws.cell(row=start_row + idx, column=13, value=str(row['cus_tel'])).border = thin_border
    ws.cell(row=start_row + idx, column=15, value=str(row['cus_branch'])).border = thin_border
    ws.cell(row=start_row + idx, column=16, value=str(row['ct_name'])).border = thin_border
    ws.cell(row=start_row + idx, column=20, value=str(row['cus_tax_no'])).border = thin_border
    ws.cell(row=start_row + idx, column=21, value=str(row['ca_detail_address'])).border = thin_border
    ws.cell(row=start_row + idx, column=23, value=str(row['ca_loc_province'])).border = thin_border
    ws.cell(row=start_row + idx, column=24, value=str(row['ca_loc_district'])).border = thin_border
    ws.cell(row=start_row + idx, column=25, value=str(row['ca_loc_data'])).border = thin_border
    ws.cell(row=start_row + idx, column=27, value=str(row['ca_postcode'])).border = thin_border
    ws.cell(row=start_row + idx, column=28, value=str(row['ca_loc_province_name'])).border = thin_border
    ws.cell(row=start_row + idx, column=29, value=str(row['ca_loc_district_name'])).border = thin_border
    ws.cell(row=start_row + idx, column=30, value=str(row['ca_loc_sub_district_name'])).border = thin_border
    ws.cell(row=start_row + idx, column=31, value=str(row['ca_postcode'])).border = thin_border
    ws.cell(row=start_row + idx, column=32, value=str(row['ca_latitude'])).border = thin_border
    ws.cell(row=start_row + idx, column=33, value=str(row['ca_longitude'])).border = thin_border
    ws.cell(row=start_row + idx, column=34, value="TRUE").border = thin_border
    ws.cell(row=start_row + idx, column=35, value="TRUE").border = thin_border
    ws.cell(row=start_row + idx, column=36, value=cus_created_date).border = thin_border
    ws.cell(row=start_row + idx, column=38, value=cus_updated_date).border = thin_border

    if ws.cell(row=start_row + idx, column=9).value == 'Delivery': # juristic_type
        ws.cell(row=start_row + idx, column=9, value="S").border = thin_border 
    elif ws.cell(row=start_row + idx, column=9).value == 'Invoice':
        ws.cell(row=start_row + idx, column=9, value="T").border = thin_border
    else:
        ws.cell(row=start_row + idx, column=11, value="").border = thin_border 
        
    if ws.cell(row=start_row + idx, column=15).value == '00000': # organize_type
        ws.cell(row=start_row + idx, column=15, value="HEAD_OFFICE").border = thin_border 
    elif ws.cell(row=start_row + idx, column=15).value == '':
        ws.cell(row=start_row + idx, column=15, value="").border = thin_border
    else:
        ws.cell(row=start_row + idx, column=11, value="BRANCH").border = thin_border

    if ws.cell(row=start_row + idx, column=16).value == 'Organization':
        ws.cell(row=start_row + idx, column=16, value="LIMITED_COMPANY").border = thin_border # juristic_type
    elif ws.cell(row=start_row + idx, column=16).value == 'Person':
        ws.cell(row=start_row + idx, column=16, value="Personal").border = thin_border
    else:
        ws.cell(row=start_row + idx, column=16, value="").border = thin_border 
        
    if ws.cell(row=start_row + idx, column=13).value == 'None': # contact_phone_number
        ws.cell(row=start_row + idx, column=13, value='').border = thin_border 
    
# เพิ่มเส้นกรอบให้กับเซลล์ที่ไม่ได้เติมค่า
for row in range(start_row, start_row + len(customer_df)):
    for col in range(1, 41):
        cell = ws.cell(row=row, column=col)
        if cell.value is None:
            cell.border = thin_border

# บันทึกไฟล์ Excel
wb.save(file_path)
print(f'บันทึกไฟล์ Excel ที่: {file_path} เสร็จสิ้น พร้อมข้อมูลจากฐานข้อมูล')
