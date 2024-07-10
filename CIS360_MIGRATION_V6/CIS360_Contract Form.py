import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side
from sqlalchemy import create_engine
from datetime import datetime

# ข้อมูลการเชื่อมต่อ
DB_USERNAME = "pondttt"
DB_PASSWORD = "2aJ998we~"
DB_NAME = "Customer_UAT"
DB_HOST = "pgsql-sku-az-sea-dev.postgres.database.azure.com"
DB_PORT = "5432"

# กำหนดการจัดวางข้อความตรงกลาง
alignment_center = Alignment(horizontal='center', vertical='center')

# กำหนดกรอบของข้อมูล
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# โหลดไฟล์ Excel ที่มีอยู่
file_path = r'E:/CIS/CIS360 Migration Form V6.xlsx'
wb = load_workbook(file_path)
ws = wb['Contact Form']

# สร้างการเชื่อมต่อด้วย SQLAlchemy
connection_string = f"postgresql://{DB_USERNAME}:{DB_PASSWORD}@{DB_HOST}:{DB_PORT}/{DB_NAME}"
engine = create_engine(connection_string)

# สร้าง DataFrame จากฐานข้อมูล
query = """
SELECT * FROM contact_form_v6
"""

customer_df = pd.read_sql(query, engine)

# เพิ่มข้อมูลจากฐานข้อมูลลงใน Excel ตั้งแต่แถวที่ 9
start_row = 9
for idx, row in customer_df.iterrows():
    cus_created_date = row['cus_created_date'].replace(tzinfo=None)
    cus_updated_date = row['cus_updated_date'].replace(tzinfo=None)
    
    ws.cell(row=start_row + idx, column=3, value="ALLKONS_AMRP").border = thin_border
    ws.cell(row=start_row + idx, column=4, value="ALLKONS_AMRP").border = thin_border
    ws.cell(row=start_row + idx, column=5, value="OFFICE").border = thin_border
    ws.cell(row=start_row + idx, column=6, value=str(row['cus_id'])).border = thin_border
    ws.cell(row=start_row + idx, column=7, value="PHONE").border = thin_border
    ws.cell(row=start_row + idx, column=9, value=str(row['cus_tel'])).border = thin_border
    ws.cell(row=start_row + idx, column=15, value="TRUE").border = thin_border
    ws.cell(row=start_row + idx, column=16, value="TRUE").border = thin_border
    ws.cell(row=start_row + idx, column=17, value="TRUE").border = thin_border
    ws.cell(row=start_row + idx, column=18, value=str(row['cus_created_date'])).border = thin_border
    ws.cell(row=start_row + idx, column=20, value=str(row['cus_updated_date'])).border = thin_border
    
# เพิ่มเส้นกรอบให้กับเซลล์ที่ไม่ได้เติมค่า
for row in range(start_row, start_row + len(customer_df)):
    for col in range(1, 23):
        cell = ws.cell(row=row, column=col)
        if cell.value is None:
            cell.border = thin_border

# บันทึกไฟล์ Excel
wb.save(file_path)
print(f'บันทึกไฟล์ Excel ที่: {file_path} เสร็จสิ้น พร้อมข้อมูลจากฐานข้อมูล')