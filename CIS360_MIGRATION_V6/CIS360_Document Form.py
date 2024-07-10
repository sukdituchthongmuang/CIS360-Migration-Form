import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side
from sqlalchemy import create_engine
from datetime import datetime

# ข้อมูลการเชื่อมต่อ
DB_USERNAME = "pondttt"
DB_PASSWORD = "2aJ998we~"
DB_NAME = "Customer"
DB_HOST = "pgsql-sku-az-sea-dev.postgres.database.azure.com"
DB_PORT = "5432"

# กำหนดการจัดวางข้อความตรงกลาง
alignment_center = Alignment(horizontal='center', vertical='center')

# กำหนดกรอบของข้อมูล
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# โหลดไฟล์ Excel ที่มีอยู่
file_path = r'E:/CIS/CIS360 Migration Form V6.xlsx'
wb = load_workbook(file_path)
ws = wb['Document Form']

# สร้างการเชื่อมต่อด้วย SQLAlchemy
connection_string = f"postgresql://{DB_USERNAME}:{DB_PASSWORD}@{DB_HOST}:{DB_PORT}/{DB_NAME}"
engine = create_engine(connection_string)

# สร้าง DataFrame จากฐานข้อมูล
query = """
SELECT * FROM document_form_v6
"""

upload_customer_df = pd.read_sql(query, engine)

# เลือก 10% ของข้อมูล
num_rows = int(len(upload_customer_df) * 0.1)
customer_df = upload_customer_df.head(num_rows).reset_index(drop=True)

start_row = 9
for idx, row in upload_customer_df.iterrows():
    uc_created_date = row['uc_created_date'].replace(tzinfo=None)
    uc_updated_date = row['uc_updated_date'].replace(tzinfo=None)
    uc_path = str(row['uc_path'])
    
    # ตัดข้อมูลที่อยู่หลังสุดจนถึงจุดแรก
    main_path, _, remainder = uc_path.rpartition('.')
    
    # ตัดข้อมูลที่อยู่หลังสุดจนถึงเครื่องหมาย /
    main_path2, _, file_name = main_path.rpartition('/')
    
    ws.cell(row=start_row + idx, column=3, value="ALLKONS_AMRP").border = thin_border
    ws.cell(row=start_row + idx, column=4, value="ALLKONS_AMRP").border = thin_border
    ws.cell(row=start_row + idx, column=5, value="OFFICE").border = thin_border
    ws.cell(row=start_row + idx, column=6, value=str(row['uc_cus_id'])).border = thin_border
    ws.cell(row=start_row + idx, column=7, value="IMAGE_PROFILE").border = thin_border
    ws.cell(row=start_row + idx, column=8, value=main_path2).border = thin_border
    ws.cell(row=start_row + idx, column=9, value=file_name).border = thin_border
    ws.cell(row=start_row + idx, column=10, value="."+remainder).border = thin_border
    ws.cell(row=start_row + idx, column=12, value=str(row['uc_is_use'])).border = thin_border
    ws.cell(row=start_row + idx, column=13, value=str(row['uc_created_date'])).border = thin_border
    ws.cell(row=start_row + idx, column=15, value=str(row['uc_updated_date'])).border = thin_border

# เพิ่มเส้นกรอบให้กับเซลล์ที่ไม่ได้เติมค่า
for row in range(start_row, start_row + len(upload_customer_df)):
    for col in range(1, 18):
        cell = ws.cell(row=row, column=col)
        if cell.value is None:
            cell.border = thin_border

# บันทึกไฟล์ Excel
wb.save(file_path)
print(f'บันทึกไฟล์ Excel ที่: {file_path} เสร็จสิ้น พร้อมข้อมูลจากฐานข้อมูล')
